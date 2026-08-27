# data_analyzer.py - Analisis Data Perangkat Pembelajaran
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import mysql.connector
from datetime import datetime, timedelta
import os
from dotenv import load_dotenv

load_dotenv()

# ============================================
# KONEKSI DATABASE
# ============================================
def get_db_connection():
    return mysql.connector.connect(
        host=os.getenv('DB_HOST', 'localhost'),
        user=os.getenv('DB_USER', 'root'),
        password=os.getenv('DB_PASSWORD', ''),
        database=os.getenv('DB_NAME', 'perangkat_pembelajaran')
    )

# ============================================
# ANALISIS DOKUMEN PER JENIS
# ============================================
def analisis_dokumen_per_jenis():
    """Analisis jumlah dokumen per jenis"""
    conn = get_db_connection()
    query = """
        SELECT 
            jenis,
            COUNT(*) as total,
            SUM(CASE WHEN status = 'terverifikasi' THEN 1 ELSE 0 END) as terverifikasi,
            SUM(CASE WHEN status = 'pending_kepsek' OR status = 'pending_pengawas' THEN 1 ELSE 0 END) as pending,
            SUM(CASE WHEN status = 'ditolak_kepsek' OR status = 'ditolak_pengawas' THEN 1 ELSE 0 END) as ditolak
        FROM dokumen_perangkat
        GROUP BY jenis
        ORDER BY total DESC
    """
    df = pd.read_sql(query, conn)
    conn.close()
    return df.to_dict('records')

# ============================================
# ANALISIS DOKUMEN PER SEKOLAH
# ============================================
def analisis_dokumen_per_sekolah():
    """Analisis dokumen per sekolah"""
    conn = get_db_connection()
    query = """
        SELECT 
            s.nama_sekolah,
            COUNT(d.id) as total_dokumen,
            SUM(CASE WHEN d.status = 'terverifikasi' THEN 1 ELSE 0 END) as terverifikasi,
            COUNT(DISTINCT d.id_guru) as total_guru_upload
        FROM sekolah s
        LEFT JOIN dokumen_perangkat d ON s.id = d.id_sekolah
        GROUP BY s.id
        HAVING total_dokumen > 0
        ORDER BY total_dokumen DESC
        LIMIT 10
    """
    df = pd.read_sql(query, conn)
    conn.close()
    return df.to_dict('records')

# ============================================
# ANALISIS TREND BULANAN
# ============================================
def analisis_trend_bulanan():
    """Analisis trend upload dokumen per bulan"""
    conn = get_db_connection()
    query = """
        SELECT 
            DATE_FORMAT(created_at, '%Y-%m') as bulan,
            COUNT(*) as total,
            SUM(CASE WHEN status = 'terverifikasi' THEN 1 ELSE 0 END) as terverifikasi
        FROM dokumen_perangkat
        WHERE created_at >= DATE_SUB(NOW(), INTERVAL 6 MONTH)
        GROUP BY DATE_FORMAT(created_at, '%Y-%m')
        ORDER BY bulan
    """
    df = pd.read_sql(query, conn)
    conn.close()
    return df.to_dict('records')

# ============================================
# ANALISIS KINERJA GURU
# ============================================
def analisis_kinerja_guru():
    """Analisis kinerja guru berdasarkan jumlah dokumen"""
    conn = get_db_connection()
    query = """
        SELECT 
            u.name as guru_name,
            COUNT(d.id) as total_dokumen,
            SUM(CASE WHEN d.status = 'terverifikasi' THEN 1 ELSE 0 END) as terverifikasi,
            SUM(CASE WHEN d.status = 'pending_kepsek' OR d.status = 'pending_pengawas' THEN 1 ELSE 0 END) as pending,
            SUM(d.views) as total_views,
            SUM(d.downloads) as total_downloads
        FROM users u
        LEFT JOIN dokumen_perangkat d ON u.id = d.id_guru
        WHERE u.role = 'guru'
        GROUP BY u.id
        HAVING total_dokumen > 0
        ORDER BY total_dokumen DESC
        LIMIT 10
    """
    df = pd.read_sql(query, conn)
    conn.close()
    return df.to_dict('records')

# ============================================
# GENERATE LAPORAN EXCEL
# ============================================
def generate_laporan_excel():
    """Generate laporan lengkap ke Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    
    # Buat workbook
    wb = Workbook()
    
    # 1. Sheet Ringkasan
    ws1 = wb.active
    ws1.title = "Ringkasan"
    
    # Data
    dokumen_per_jenis = analisis_dokumen_per_jenis()
    trend_bulanan = analisis_trend_bulanan()
    
    # Header
    ws1['A1'] = "LAPORAN PERANGKAT PEMBELAJARAN"
    ws1.merge_cells('A1:F1')
    ws1['A1'].font = Font(size=16, bold=True)
    
    ws1['A3'] = "Tanggal"
    ws1['B3'] = datetime.now().strftime('%d/%m/%Y %H:%M')
    
    # Statistik per Jenis
    ws1['A5'] = "STATISTIK PER JENIS"
    ws1['A5'].font = Font(bold=True)
    
    headers = ['Jenis', 'Total', 'Terverifikasi', 'Pending', 'Ditolak']
    for col, header in enumerate(headers, 1):
        ws1.cell(row=6, column=col, value=header).font = Font(bold=True)
        ws1.cell(row=6, column=col).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws1.cell(row=6, column=col).font = Font(color="FFFFFF", bold=True)
    
    for row, data in enumerate(dokumen_per_jenis, 7):
        ws1.cell(row=row, column=1, value=data['jenis'])
        ws1.cell(row=row, column=2, value=data['total'])
        ws1.cell(row=row, column=3, value=data['terverifikasi'])
        ws1.cell(row=row, column=4, value=data['pending'])
        ws1.cell(row=row, column=5, value=data['ditolak'])
    
    # Trend Bulanan
    ws1['A' + str(len(dokumen_per_jenis) + 9)] = "TREND BULANAN"
    ws1['A' + str(len(dokumen_per_jenis) + 9)].font = Font(bold=True)
    
    ws1.cell(row=len(dokumen_per_jenis) + 10, column=1, value="Bulan").font = Font(bold=True)
    ws1.cell(row=len(dokumen_per_jenis) + 10, column=2, value="Total").font = Font(bold=True)
    ws1.cell(row=len(dokumen_per_jenis) + 10, column=3, value="Terverifikasi").font = Font(bold=True)
    
    for row, data in enumerate(trend_bulanan, len(dokumen_per_jenis) + 11):
        ws1.cell(row=row, column=1, value=data['bulan'])
        ws1.cell(row=row, column=2, value=data['total'])
        ws1.cell(row=row, column=3, value=data['terverifikasi'])
    
    # 2. Sheet Kinerja Guru
    ws2 = wb.create_sheet("Kinerja Guru")
    
    kinerja = analisis_kinerja_guru()
    headers = ['No', 'Nama Guru', 'Total Dokumen', 'Terverifikasi', 'Pending', 'Views', 'Downloads']
    for col, header in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=header).font = Font(bold=True)
        ws2.cell(row=1, column=col).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws2.cell(row=1, column=col).font = Font(color="FFFFFF", bold=True)
    
    for row, data in enumerate(kinerja, 2):
        ws2.cell(row=row, column=1, value=row-1)
        ws2.cell(row=row, column=2, value=data['guru_name'])
        ws2.cell(row=row, column=3, value=data['total_dokumen'])
        ws2.cell(row=row, column=4, value=data['terverifikasi'])
        ws2.cell(row=row, column=5, value=data['pending'])
        ws2.cell(row=row, column=6, value=data['total_views'])
        ws2.cell(row=row, column=7, value=data['total_downloads'])
    
    # 3. Sheet Sekolah
    ws3 = wb.create_sheet("Sekolah")
    
    sekolah = analisis_dokumen_per_sekolah()
    headers = ['No', 'Nama Sekolah', 'Total Dokumen', 'Terverifikasi', 'Total Guru Upload']
    for col, header in enumerate(headers, 1):
        ws3.cell(row=1, column=col, value=header).font = Font(bold=True)
        ws3.cell(row=1, column=col).fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        ws3.cell(row=1, column=col).font = Font(color="FFFFFF", bold=True)
    
    for row, data in enumerate(sekolah, 2):
        ws3.cell(row=row, column=1, value=row-1)
        ws3.cell(row=row, column=2, value=data['nama_sekolah'])
        ws3.cell(row=row, column=3, value=data['total_dokumen'])
        ws3.cell(row=row, column=4, value=data['terverifikasi'])
        ws3.cell(row=row, column=5, value=data['total_guru_upload'])
    
    # Simpan ke bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ============================================
# GENERATE VISUALISASI
# ============================================
def generate_visualisasi():
    """Generate visualisasi data"""
    # Set style
    plt.style.use('seaborn-v0_8-darkgrid')
    sns.set_palette("husl")
    
    fig, axes = plt.subplots(2, 2, figsize=(14, 10))
    fig.suptitle('Dashboard Analisis Perangkat Pembelajaran', fontsize=16, fontweight='bold')
    
    # 1. Dokumen per Jenis
    data = analisis_dokumen_per_jenis()
    if data:
        df = pd.DataFrame(data)
        ax = axes[0, 0]
        bars = ax.bar(df['jenis'], df['total'], color=sns.color_palette("husl", len(df)))
        ax.set_title('Jumlah Dokumen per Jenis', fontsize=12, fontweight='bold')
        ax.set_xlabel('Jenis')
        ax.set_ylabel('Jumlah')
        ax.tick_params(axis='x', rotation=45)
        # Tambahkan label nilai
        for bar, val in zip(bars, df['total']):
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5, str(val), ha='center', va='bottom', fontsize=10)
    
    # 2. Status Dokumen
    if data:
        df = pd.DataFrame(data)
        ax = axes[0, 1]
        status_data = {
            'Terverifikasi': df['terverifikasi'].sum(),
            'Pending': df['pending'].sum(),
            'Ditolak': df['ditolak'].sum()
        }
        colors = ['#28a745', '#ffc107', '#dc3545']
        wedges, texts, autotexts = ax.pie(
            status_data.values(), 
            labels=status_data.keys(),
            autopct='%1.1f%%',
            colors=colors,
            startangle=90
        )
        ax.set_title('Status Dokumen', fontsize=12, fontweight='bold')
        for autotext in autotexts:
            autotext.set_color('white')
            autotext.set_fontweight('bold')
    
    # 3. Trend Bulanan
    trend = analisis_trend_bulanan()
    if trend:
        df = pd.DataFrame(trend)
        ax = axes[1, 0]
        ax.plot(df['bulan'], df['total'], marker='o', linewidth=2, label='Total')
        ax.plot(df['bulan'], df['terverifikasi'], marker='s', linewidth=2, label='Terverifikasi')
        ax.set_title('Trend Upload per Bulan', fontsize=12, fontweight='bold')
        ax.set_xlabel('Bulan')
        ax.set_ylabel('Jumlah')
        ax.legend()
        ax.tick_params(axis='x', rotation=45)
    
    # 4. Kinerja Guru (Top 5)
    kinerja = analisis_kinerja_guru()
    if kinerja:
        df = pd.DataFrame(kinerja[:5])
        ax = axes[1, 1]
        bars = ax.barh(df['guru_name'], df['total_dokumen'], color=sns.color_palette("husl", len(df)))
        ax.set_title('Top 5 Guru dengan Dokumen Terbanyak', fontsize=12, fontweight='bold')
        ax.set_xlabel('Jumlah Dokumen')
        ax.set_ylabel('Nama Guru')
        # Tambahkan label nilai
        for bar, val in zip(bars, df['total_dokumen']):
            ax.text(bar.get_width() + 0.5, bar.get_y() + bar.get_height()/2, str(val), ha='left', va='center', fontsize=10)
    
    plt.tight_layout()
    
    # Simpan ke bytes
    import io
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    buf.seek(0)
    plt.close()
    return buf