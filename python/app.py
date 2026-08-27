# app.py - Flask API Sederhana (Tanpa Pandas)
from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
from datetime import datetime
import os
import io

app = Flask(__name__)
CORS(app)

# ============================================
# FUNGSI-FUNGSI
# ============================================

def generate_perangkat_ajar(jenis, topik, kelas, mapel):
    return {
        'judul': f"{jenis.upper()} {mapel} Kelas {kelas} - {topik}",
        'deskripsi': f"Dokumen {jenis} ini membahas tentang {topik} untuk kelas {kelas} mata pelajaran {mapel}.",
        'tujuan': f"1. Memahami konsep {topik}\n2. Menerapkan {topik}\n3. Menganalisis {topik}\n4. Mengevaluasi {topik}",
        'langkah': f"1. Pendahuluan (10 menit)\n2. Kegiatan Inti (60 menit)\n3. Penutup (20 menit)",
        'penilaian': "1. Penilaian Sikap\n2. Penilaian Pengetahuan\n3. Penilaian Keterampilan"
    }

def analisis_dokumen_per_jenis():
    return [
        {'jenis': 'rpp', 'total': 15, 'terverifikasi': 10, 'pending': 3, 'ditolak': 2},
        {'jenis': 'modul', 'total': 8, 'terverifikasi': 5, 'pending': 2, 'ditolak': 1},
        {'jenis': 'ppt', 'total': 12, 'terverifikasi': 8, 'pending': 3, 'ditolak': 1},
        {'jenis': 'soal', 'total': 6, 'terverifikasi': 4, 'pending': 1, 'ditolak': 1}
    ]

def analisis_trend_bulanan():
    return [
        {'bulan': '2026-03', 'total': 5, 'terverifikasi': 3},
        {'bulan': '2026-04', 'total': 8, 'terverifikasi': 5},
        {'bulan': '2026-05', 'total': 12, 'terverifikasi': 8},
        {'bulan': '2026-06', 'total': 10, 'terverifikasi': 7},
        {'bulan': '2026-07', 'total': 15, 'terverifikasi': 10},
        {'bulan': '2026-08', 'total': 20, 'terverifikasi': 14}
    ]

def generate_laporan_excel():
    """Generate laporan Excel (tanpa pandas)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan"
        
        # Header
        ws.merge_cells('A1:F1')
        ws['A1'] = "LAPORAN PERANGKAT PEMBELAJARAN"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Tanggal
        ws['A3'] = "Tanggal"
        ws['B3'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        
        # Header tabel
        headers = ['Jenis', 'Total', 'Terverifikasi', 'Pending', 'Ditolak']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Data
        data = analisis_dokumen_per_jenis()
        for row, item in enumerate(data, 6):
            ws.cell(row=row, column=1, value=item['jenis'])
            ws.cell(row=row, column=2, value=item['total'])
            ws.cell(row=row, column=3, value=item['terverifikasi'])
            ws.cell(row=row, column=4, value=item['pending'])
            ws.cell(row=row, column=5, value=item['ditolak'])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    except:
        # Jika openpyxl tidak terinstall
        return None

def generate_visualisasi():
    """Generate visualisasi (tanpa matplotlib)"""
    import base64
    # Buat chart sederhana dengan HTML
    data = analisis_dokumen_per_jenis()
    html = """
    <html>
    <head>
        <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    </head>
    <body style="padding:20px;">
        <h2>📊 Dokumen per Jenis</h2>
        <canvas id="chart" width="600" height="400"></canvas>
        <script>
            const ctx = document.getElementById('chart').getContext('2d');
            new Chart(ctx, {
                type: 'bar',
                data: {
                    labels: """ + str([d['jenis'].upper() for d in data]) + """,
                    datasets: [{
                        label: 'Total',
                        data: """ + str([d['total'] for d in data]) + """,
                        backgroundColor: ['#4f46e5', '#0d9488', '#2563eb', '#7c3aed']
                    }]
                },
                options: {
                    responsive: true,
                    plugins: { legend: { display: false } }
                }
            });
        </script>
    </body>
    </html>
    """
    return html

# ============================================
# ROUTES
# ============================================

@app.route('/')
def home():
    return render_template('dashboard.html')

@app.route('/api/generate/perangkat', methods=['POST'])
def api_generate_perangkat():
    data = request.json
    result = generate_perangkat_ajar(
        data.get('jenis', 'rpp'),
        data.get('topik', ''),
        data.get('kelas', ''),
        data.get('mapel', '')
    )
    return jsonify(result)

@app.route('/api/analisis/jenis', methods=['GET'])
def api_analisis_jenis():
    return jsonify(analisis_dokumen_per_jenis())

@app.route('/api/analisis/trend', methods=['GET'])
def api_analisis_trend():
    return jsonify(analisis_trend_bulanan())

@app.route('/api/laporan/excel', methods=['GET'])
def api_laporan_excel():
    output = generate_laporan_excel()
    if output:
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'laporan_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    else:
        return jsonify({'error': 'Openpyxl not installed'}), 500

@app.route('/api/visualisasi', methods=['GET'])
def api_visualisasi():
    return generate_visualisasi()

# ============================================
# RUN
# ============================================
if __name__ == '__main__':
    os.makedirs('templates', exist_ok=True)
    print("=" * 50)
    print("🚀 PYTHON AI SERVER STARTED")
    print("=" * 50)
    print(f"🌐 URL: http://localhost:5000")
    print("=" * 50)
    app.run(debug=True, host='0.0.0.0', port=5000)